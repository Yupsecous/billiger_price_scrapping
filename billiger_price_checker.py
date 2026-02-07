"""Billiger.de price checker — fetches lowest prices by EAN code."""

import logging
import random
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional, Tuple

import pandas as pd
import undetected_chromedriver as uc
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


class BilligerPriceChecker:
    """Checks billiger.de for the lowest product prices via undetected Chrome."""

    BASE_URL = "https://www.billiger.de"

    def __init__(
        self,
        headless: bool = False,
        delay_range: Tuple[float, float] = (1.0, 2.0),
    ):
        self.delay_range = delay_range
        self.headless = headless
        self.driver = None
        self.request_count = 0

    # -- Lifecycle ------------------------------------------------------------

    def _init_driver(self):
        """Start undetected Chrome and establish a session with billiger.de."""
        if self.driver is not None:
            return

        options = uc.ChromeOptions()
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--lang=de-DE")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--no-sandbox")
        if self.headless:
            options.add_argument("--headless=new")

        logger.info("Initializing Chrome WebDriver ...")
        self.driver = uc.Chrome(
            options=options,
            use_subprocess=True,
            version_main=144,
            driver_executable_path=None,
        )
        self.driver.implicitly_wait(10)
        logger.info("Chrome WebDriver ready")

        logger.info("Establishing session with billiger.de ...")
        self.driver.get(self.BASE_URL)
        try:
            WebDriverWait(self.driver, 10).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
        except TimeoutException:
            pass
        time.sleep(2)

        title = self.driver.title.lower()
        if "billiger.de" in title or "preisvergleich" in title:
            logger.info("Session established")
        else:
            logger.warning("Possible Cloudflare challenge — waiting longer")
            time.sleep(5)

    def close(self):
        """Shut down the browser."""
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = None
            logger.info("WebDriver closed")

    # -- Public API -----------------------------------------------------------

    def get_price(self, ean: str) -> Dict:
        """Return the lowest billiger / eBay price for *ean*."""
        self._init_driver()
        ean = str(ean).strip()
        logger.info(f"Searching EAN: {ean}")
        return self._search(ean) or self._empty_result()

    # -- Private helpers ------------------------------------------------------

    def _human_delay(self):
        """Sleep a random interval with occasional longer pauses."""
        if self.request_count and self.request_count % 100 == 0:
            logger.info(f"Short pause after {self.request_count} requests")
            time.sleep(random.uniform(3, 6))
        if self.request_count and self.request_count % 500 == 0:
            logger.info(f"Longer pause after {self.request_count} requests")
            time.sleep(random.uniform(8, 15))
        time.sleep(random.uniform(*self.delay_range))
        self.request_count += 1

    def _find_search_box(self):
        """Locate the visible search input on the current page."""
        for sel in (
            'input[name="searchstring"]',
            'input[type="search"]',
            'input[placeholder*="uche"]',
            'input[id*="search"]',
            'input[class*="search"]',
        ):
            try:
                for el in self.driver.find_elements(By.CSS_SELECTOR, sel):
                    if el.is_displayed() and el.is_enabled():
                        return el
            except Exception:
                continue
        return None

    def _type_human(self, element, text: str):
        """Send keystrokes with small random delays."""
        for ch in text:
            element.send_keys(ch)
            time.sleep(random.uniform(0.02, 0.08))

    def _is_cloudflare(self) -> bool:
        """Return True if the current page shows a Cloudflare challenge."""
        src = self.driver.page_source
        markers = (
            "Nur einen Moment",
            "Just a moment",
            "Checking your browser",
            "\u00dcberpr\u00fcfung Ihres Browsers",
        )
        return any(m in src for m in markers)

    # -- Search ---------------------------------------------------------------

    def _search(self, query: str) -> Optional[Dict]:
        """Submit *query* via the on-page search box and return results."""
        self._human_delay()
        try:
            search_box = self._find_search_box()

            if not search_box:
                logger.debug("Search box not found — navigating to homepage")
                self.driver.get(self.BASE_URL)
                try:
                    WebDriverWait(self.driver, 10).until(
                        lambda d: d.execute_script("return document.readyState") == "complete"
                    )
                except TimeoutException:
                    pass
                time.sleep(random.uniform(1.0, 2.0))
                if self._is_cloudflare():
                    logger.warning("Cloudflare challenge on homepage — waiting")
                    time.sleep(random.uniform(5, 10))
                search_box = self._find_search_box()

            if not search_box:
                logger.warning("Search box unavailable")
                return None

            search_box.click()
            time.sleep(random.uniform(0.05, 0.15))
            search_box.send_keys(Keys.CONTROL + "a")
            time.sleep(random.uniform(0.03, 0.08))
            self._type_human(search_box, query)
            time.sleep(random.uniform(0.15, 0.35))

            search_box.send_keys(Keys.RETURN)
            logger.debug(f"Submitted search: {query}")

            try:
                WebDriverWait(self.driver, 12).until(
                    lambda d: (
                        "/products/" in d.current_url
                        or "/noresult" in d.current_url
                        or d.find_elements(By.CSS_SELECTOR, 'a[href*="/products/"]')
                        or "keine produkte" in d.page_source.lower()
                    )
                )
            except TimeoutException:
                logger.warning(f"Timeout waiting for results: {query}")

            time.sleep(random.uniform(0.3, 0.7))

            if self._is_cloudflare():
                logger.warning("Cloudflare challenge on results — waiting")
                time.sleep(random.uniform(5, 10))

            url = self.driver.current_url
            if "/noresult" in url:
                logger.info("No results page")
                return None

            src = self.driver.page_source.lower()
            if "keine produkte gefunden" in src or "leider keine" in src:
                logger.info("No products found")
                return None

            if "/products/" in url:
                logger.info("Direct product page")
                return self._extract_product_page()

            return self._extract_search_results()

        except Exception as exc:
            logger.error(f"Search error: {exc}")
            return None

    # -- Result extraction ----------------------------------------------------

    _CARDS_BASE = "/html/body/div[3]/div[6]/div[4]/div[2]/div[5]/div[3]/div"

    def _extract_search_results(self) -> Optional[Dict]:
        """Parse billiger / eBay prices from product cards on the results page."""
        try:
            time.sleep(random.uniform(0.3, 0.6))
            result = self._empty_result()
            billiger, ebay = [], []

            for idx in range(1, 13):
                xpath = f"{self._CARDS_BASE}/div[{idx}]"
                try:
                    card = self.driver.find_element(By.XPATH, xpath)
                except Exception:
                    if idx > 1:
                        break
                    continue

                if self._scan_billiger_card(xpath, idx, billiger):
                    if ebay:
                        break
                    continue

                self._scan_ebay_card(card, xpath, idx, ebay)
                if billiger and ebay:
                    break

            if not billiger and not ebay:
                self._scan_single_result(billiger, ebay)

            return self._pick_lowest(result, billiger, ebay)

        except Exception as exc:
            logger.error(f"Error extracting search results: {exc}")
            return None

    def _scan_billiger_card(self, xpath: str, idx: int, prices: list) -> bool:
        for sub in (f"{xpath}/div[3]/div[2]", f"{xpath}/div[2]/div[2]"):
            try:
                text = self.driver.find_element(By.XPATH, sub).text.strip()
                if text and text.lower().startswith("ab"):
                    m = re.search(r"(\d+[.,]\d{2})", text)
                    if m:
                        p = self._parse_price(m.group(1))
                        if p and p > 0:
                            prices.append(p)
                            logger.info(f"  Card {idx}: billiger — {p} EUR")
                            return True
            except Exception:
                continue
        return False

    def _scan_ebay_card(self, card, xpath: str, idx: int, prices: list):
        is_ebay = False
        for img_path in (
            f"{xpath}/div[2]/div[2]/div[2]/img",
            f"{xpath}/div[2]/div[2]/img",
        ):
            try:
                alt = self.driver.find_element(By.XPATH, img_path).get_attribute("alt") or ""
                if "ebay" in alt.lower():
                    is_ebay = True
                    break
            except Exception:
                continue
        if not is_ebay:
            try:
                card.find_element(By.XPATH, ".//img[contains(@alt, 'ebay')]")
                is_ebay = True
            except Exception:
                pass
        if not is_ebay:
            return

        for price_path in (
            f"{xpath}/div[2]/div[2]/div[1]",
            f"{xpath}/div[2]/div[1]",
        ):
            try:
                text = self.driver.find_element(By.XPATH, price_path).text.strip()
                m = re.search(r"(\d+[.,]\d{2})", text)
                if m:
                    p = self._parse_price(m.group(1))
                    if p and p > 0:
                        prices.append(p)
                        logger.info(f"  Card {idx}: eBay — {p} EUR")
                        return
            except Exception:
                continue

    def _scan_single_result(self, billiger: list, ebay: list):
        """Handle a single non-indexed card."""
        base = self._CARDS_BASE
        try:
            el = self.driver.find_element(By.XPATH, f"{base}/div/div[3]/div[2]")
            text = el.text.strip()
            if text and text.lower().startswith("ab"):
                m = re.search(r"(\d+[.,]\d{2})", text)
                if m:
                    p = self._parse_price(m.group(1))
                    if p and p > 0:
                        billiger.append(p)
                        logger.info(f"  Single result: billiger — {p} EUR")
        except Exception:
            pass

        try:
            card = self.driver.find_element(By.XPATH, f"{base}/div")
            card.find_element(By.XPATH, ".//img[contains(@alt, 'ebay')]")
            for pp in (
                f"{base}/div/div[2]/div[2]/div[1]",
                f"{base}/div/div[2]/div[1]",
            ):
                try:
                    text = self.driver.find_element(By.XPATH, pp).text.strip()
                    m = re.search(r"(\d+[.,]\d{2})", text)
                    if m:
                        p = self._parse_price(m.group(1))
                        if p and p > 0:
                            ebay.append(p)
                            logger.info(f"  Single result: eBay — {p} EUR")
                            return
                except Exception:
                    continue
        except Exception:
            pass

    def _extract_product_page(self) -> Optional[Dict]:
        """Parse billiger / eBay prices from a product detail page."""
        result = self._empty_result()
        try:
            time.sleep(random.uniform(0.3, 0.6))
            body_text = self.driver.find_element(By.TAG_NAME, "body").text
            billiger, ebay = [], []

            for selector in (
                '[class*="offer"]',
                '[class*="shop"]',
                '[class*="merchant"]',
                '[class*="dealer"]',
                'tr[class*="row"]',
                '[class*="price-list"] > div',
                '[class*="pricelist"] > div',
            ):
                try:
                    offers = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    if not offers:
                        continue
                    for offer in offers:
                        try:
                            txt = offer.text.lower()
                            html = offer.get_attribute("innerHTML").lower()
                            is_ebay = "ebay" in txt or "ebay" in html
                            is_billiger = ("billiger" in txt or "billiger" in html) and not is_ebay
                            if not is_ebay and not is_billiger:
                                continue
                            for pm in re.findall(r"(\d+[.,]\d{2})\s*\u20ac", offer.text):
                                p = self._parse_price(pm)
                                if p and 0 < p < 50_000:
                                    (ebay if is_ebay else billiger).append(p)
                                    break
                        except Exception:
                            continue
                    if billiger or ebay:
                        break
                except Exception:
                    continue

            if not billiger and not ebay:
                excluded = {"monat", "mieten", "versand", "shipping", "lieferung"}
                lines = body_text.split("\n")
                for i, line in enumerate(lines):
                    low = line.lower()
                    if any(t in low for t in excluded):
                        continue
                    ctx = " ".join(lines[max(0, i - 1):i + 2]).lower()
                    is_ebay = "ebay" in ctx
                    is_billiger = "billiger" in ctx and not is_ebay
                    if not is_ebay and not is_billiger:
                        continue
                    for pm in re.findall(r"(\d+[.,]\d{2})\s*\u20ac", line):
                        p = self._parse_price(pm)
                        if p and 0 < p < 50_000:
                            (ebay if is_ebay else billiger).append(p)

            return self._pick_lowest(result, billiger, ebay)

        except Exception as exc:
            logger.error(f"Error extracting product page: {exc}")
            return None

    # -- Utilities ------------------------------------------------------------

    @staticmethod
    def _empty_result() -> Dict:
        return {
            "billiger_price": None,
            "ebay_price": None,
            "source": None,
            "status": "NOT FOUND",
        }

    @staticmethod
    def _pick_lowest(result: Dict, billiger: list, ebay: list) -> Dict:
        candidates = []
        if billiger:
            candidates.append(("billiger", min(billiger)))
            logger.info(f"  Lowest billiger: {min(billiger)} EUR ({len(billiger)} offers)")
        if ebay:
            candidates.append(("ebay", min(ebay)))
            logger.info(f"  Lowest eBay: {min(ebay)} EUR ({len(ebay)} offers)")
        if candidates:
            candidates.sort(key=lambda x: x[1])
            source, price = candidates[0]
            key = "billiger_price" if source == "billiger" else "ebay_price"
            result[key] = price
            result["source"] = source
            result["status"] = "Found"
            logger.info(f"  \u2192 Best: {price} EUR from {source}")
        return result

    @staticmethod
    def _parse_price(text: str) -> Optional[float]:
        """Convert a German-formatted price string to a float."""
        if not text:
            return None
        try:
            cleaned = re.sub(r"[\u20ac$\u00a3\s*ab]", "", str(text).strip())
            if "," in cleaned and "." in cleaned:
                cleaned = cleaned.replace(".", "").replace(",", ".")
            elif "," in cleaned:
                cleaned = cleaned.replace(",", ".")
            m = re.search(r"(\d+\.?\d*)", cleaned)
            return float(m.group(1)) if m else None
        except (ValueError, AttributeError):
            return None


# ---------------------------------------------------------------------------
# Excel processing
# ---------------------------------------------------------------------------

def process_excel_file(
    input_file: str,
    ean_column: str = None,
    start_row: int = 0,
    limit: int = 0,
    save_interval: int = 10,
    headless: bool = False,
) -> str:
    """Process an Excel file of EANs and write billiger.de prices back."""
    filepath = Path(input_file)
    if not filepath.exists():
        raise FileNotFoundError(f"File not found: {input_file}")

    logger.info(f"Reading {input_file}")
    df = pd.read_excel(input_file, dtype=str)
    total = len(df)
    logger.info(f"Total rows: {total}")

    if ean_column is None:
        for col in df.columns:
            if any(k in col.lower() for k in ("ean", "gtin", "barcode")):
                ean_column = col
                break
        if ean_column is None:
            ean_column = df.columns[0]
    logger.info(f"EAN column: {ean_column}")

    to_drop = [c for c in df.columns if "unnamed" in c.lower() or c.strip() == ""]
    if to_drop:
        df.drop(columns=to_drop, inplace=True)

    for col in ("billiger", "eBay", "Timestamp", "Status"):
        if col not in df.columns:
            df[col] = None

    price_idx = next(
        (i for i, c in enumerate(df.columns) if "price" in c.lower()), None
    )
    if price_idx is not None:
        cols = [c for c in df.columns if c not in ("billiger", "eBay", "Timestamp", "Status")]
        for j, nc in enumerate(("billiger", "eBay", "Timestamp", "Status")):
            cols.insert(price_idx + 1 + j, nc)
        df = df[cols]

    end_row = min(start_row + limit, total) if limit > 0 else total
    logger.info(f"Processing rows {start_row + 1} to {end_row}")

    checker = BilligerPriceChecker(headless=headless, delay_range=(1.0, 2.0))
    found, processed = 0, 0

    try:
        for idx in range(start_row, end_row):
            ean = str(df.iloc[idx][ean_column]).strip() if pd.notna(df.iloc[idx][ean_column]) else ""

            if not ean or ean.lower() == "nan" or len(ean) < 8:
                df.at[idx, "Status"] = "Invalid EAN"
                df.at[idx, "Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                processed += 1
                continue

            current = df.at[idx, "Status"]
            if pd.notna(current) and current not in ("", "None"):
                continue

            logger.info(f"[{idx + 1}/{end_row}] EAN: {ean}")
            res = checker.get_price(ean)

            df.at[idx, "billiger"] = res.get("billiger_price")
            df.at[idx, "eBay"] = res.get("ebay_price")
            df.at[idx, "Status"] = res["status"]
            df.at[idx, "Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            processed += 1

            if res["status"] == "Found":
                found += 1
                price = res.get("billiger_price") or res.get("ebay_price")
                logger.info(f"  \u2192 {price} EUR ({res.get('source', '')})")
            else:
                logger.info("  \u2192 NOT FOUND")

            if processed % save_interval == 0:
                out = filepath.parent / f"{filepath.stem}_output{filepath.suffix}"
                _save_excel(df, str(out))
                logger.info(f"Progress saved ({found}/{processed} found)")

    except KeyboardInterrupt:
        logger.info("Interrupted — saving progress")
    finally:
        checker.close()

    out = filepath.parent / f"{filepath.stem}_output{filepath.suffix}"
    _save_excel(df, str(out))

    logger.info("=" * 50)
    logger.info(f"Done  |  processed: {processed}  |  found: {found}  |  missed: {processed - found}")
    logger.info(f"Output: {out}")
    logger.info("=" * 50)
    return str(out)


def _save_excel(df: pd.DataFrame, filepath: str):
    """Write *df* to an Excel file with professional formatting."""
    df.to_excel(filepath, index=False)
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        headers = [c.value for c in ws[1]]

        col_idx = {}
        for name in ("billiger", "eBay", "Status", "Timestamp"):
            if name in headers:
                col_idx[name] = headers.index(name) + 1

        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        for name in ("billiger", "eBay"):
            if name in col_idx:
                ci = col_idx[name]
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=ci)
                    if cell.value and pd.notna(cell.value):
                        cell.number_format = '#,##0.00 "\u20ac"'
                        cell.alignment = Alignment(horizontal="right")

        if "Status" in col_idx:
            si = col_idx["Status"]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=si)
                cell.alignment = Alignment(horizontal="center")
                val = str(cell.value) if cell.value else ""
                if val == "Found":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100")
                elif "NOT FOUND" in val:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006")

        for col in ws.columns:
            letter = col[0].column_letter
            width = max((len(str(c.value or "")) for c in col), default=8) + 3
            ws.column_dimensions[letter].width = min(width, 50)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(filepath)
    except Exception as exc:
        logger.warning(f"Could not apply formatting: {exc}")


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Fetch billiger.de prices by EAN")
    parser.add_argument("input_file", help="Excel file with EAN codes")
    parser.add_argument("--ean-column", help="Name of EAN column (auto-detected)")
    parser.add_argument("--start", type=int, default=0, help="Start row (0-indexed)")
    parser.add_argument("--limit", type=int, default=0, help="Max rows (0 = all)")
    parser.add_argument("--save-interval", type=int, default=10, help="Save every N rows")
    parser.add_argument("--headless", action="store_true", help="Headless mode")
    args = parser.parse_args()

    process_excel_file(
        input_file=args.input_file,
        ean_column=args.ean_column,
        start_row=args.start,
        limit=args.limit,
        save_interval=args.save_interval,
        headless=args.headless,
    )


if __name__ == "__main__":
    main()
